"""
Program that will write cassandra table in excel workbook

"""
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from cassandra.cluster import Cluster
from cassandra.auth import PlainTextAuthProvider
from cassandra.query import SimpleStatement


dir_excel='C:\\quartExcel\\cassandraThesis.xlsx'
current_dir = os.getcwd()

def main():
    dexist=False
    dexist=os.path.exists('C:\\quartExcel')
    if dexist==False:
        os.mkdir('C:\\quartExcel')
    wb = Workbook()
    ws = wb.active
    ws.title = "Thesis" 
    #Get cassandra columns
    query="select column_name from system_schema.columns WHERE keyspace_name = 'thesis' AND table_name = 'tbthesis';"
    columns_list=''
    columns_list=cassandraBDProcess(True,query)
    coln=1
    for col in columns_list:
        #Write(row,column)
        #Headers (h1,...)
        h1 = ws.cell(row = 1, column = coln)
        h1.value = col[0]
        coln=coln+1
        wb.save(dir_excel) 
    #Starts  the reading of periods    
    for i in range(10,11):
        flag=os.path.isfile(dir_excel)
        if flag:
            #Expedient xls already exists
            print('Printing period number:',str(i))
            resultSet=''
            query="select book_number,dt_publication_date,heading,id_thesis,instance,jurisprudence_type,lst_precedents,multiple_subjects,page,period,period_number,publication,publication_date,source,subject,subject_1,subject_2,subject_3,text_content,thesis_number,type_of_thesis from thesis.tbthesis where period_number="+str(i)+""
            resultSet=cassandraBDProcess(False,query)

    print('The excel is ready!')        
                         
            
                     

def cassandraBDProcess(isShortQuery,query):

    global current_dir
    #Connect to Cassandra
    objCC=CassandraConnection()
    cloud_config= {
        'secure_connect_bundle':current_dir+'\\secure-connect-dbquart.zip'
    }
    
    auth_provider = PlainTextAuthProvider(objCC.cc_user,objCC.cc_pwd)
    res=''

    if isShortQuery:
        
        cluster = Cluster(cloud=cloud_config, auth_provider=auth_provider)
        session = cluster.connect()
        session.default_timeout=70
        #Check wheter or not the record exists      
        future = session.execute_async(query)
        res=future.result()
        cluster.shutdown()     
                
                
    else:

        cluster = Cluster(cloud=cloud_config, auth_provider=auth_provider)
        session = cluster.connect()
        session.default_timeout=70      
        statement = SimpleStatement(query, fetch_size=1000)
        wb = load_workbook(dir_excel)
        ws = wb['Thesis']
        
        for row in session.execute(statement):
            ls=[]
            for r in row:
               ls.append(str(r))
            ws.append(ls)
        
        wb.save(dir_excel) 
        cluster.shutdown() 
        res=''
            
               
                         
    return res

        


    
class CassandraConnection():
    cc_user='quartadmin'
    cc_keyspace='thesis'
    cc_pwd='P@ssw0rd33'
    cc_databaseID='9de16523-0e36-4ff0-b388-44e8d0b1581f'

if __name__=='__main__':
    main()    
   